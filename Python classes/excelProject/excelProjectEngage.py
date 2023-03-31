import openpyxl
wb = openpyxl.load_workbook("excelProjectDB.xlsx")

def getIndex(id, sheet):
    ws = wb[sheet]

    for i in range(1,ws.max_row+1):
        if ws.cell(row=i,column=1).value == id:
            return i

    return -1

#Classes
class trainee:
    def __init__(self, id, name, degree, workExperience):
        self.id =id
        self.name = name
        self.degree = degree
        self.workExperience = workExperience

    def addTrainee(self):
        ws = wb["ListOfTrainees"]
        new_row = (self.id, self.name, self.degree, self.workExperience)
        ws.append(new_row)
        wb.save("excelProjectDB.xlsx")

    def updateTrainee(self):
        index = getIndex(self.id, "ListOfTrainees")
        ws = wb["ListOfTrainees"]
        ws.cell(row=index, column=2, value = self.name)
        ws.cell(row=index, column=3, value=self.degree)
        ws.cell(row=index, column=4, value=self.workExperience)
        wb.save("excelProjectDB.xlsx")

    def printName(self):
        print(self.id, self.name)

    def printDetails(self):
        print(self.id, self.name, self.degree, self.workExperience)

class course:
    def __init__(self, id, description):
        self.id=id
        self.description = description

    def addCourse(self):
        ws = wb["CourseDetails"]
        new_row = (self.id, self.description)
        ws.append(new_row)
        wb.save("excelProjectDB.xlsx")

    def updateCourse(self):
        index = getIndex(self.id, "CourseDetails")
        ws = wb["CourseDetails"]
        ws.cell(row=index, column=2, value = self.description)
        wb.save("excelProjectDB.xlsx")

    def printDetails(self):
        print(self.id, self.description)

class trainer:
    def __init__(self, emailID, fullName, phoneNum):
        self.emailID = emailID
        self.fullName = fullName
        self.phoneNum = phoneNum

    def addTrainer(self):
        ws = wb["TrainerDetails"]
        new_row = (self.emailID, self.fullName, self.phoneNum)
        ws.append(new_row)
        wb.save("excelProjectDB.xlsx")

    def updateTrainer(self):
        index = getIndex(self.emailID, "TrainerDetails")
        ws = wb["TrainerDetails"]
        ws.cell(row=index, column=2, value=self.fullName)
        ws.cell(row=index, column=3, value=self.phoneNum)
        wb.save("excelProjectDB.xlsx")

    def printName(self):
        print(self.emailID, self.fullName)

    def printDetails(self):
        print(self.emailID, self.fullName, self.phoneNum)

class manager():
    def __init__(self, emailID, fullName, phoneNum):
        self.emailID = emailID
        self.fullName = fullName
        self.phoneNum = phoneNum

    def addManager(self):
        ws = wb["ManagerDetails"]
        new_row = (self.emailID, self.fullName, self.phoneNum)
        ws.append(new_row)
        wb.save("excelProjectDB.xlsx")

    def updateManager(self):
        index = getIndex(self.emailID, "ManagerDetails")
        ws = wb["ManagerDetails"]
        ws.cell(row=index, column=2, value=self.fullName)
        ws.cell(row=index, column=3, value=self.phoneNum)
        wb.save("excelProjectDB.xlsx")

    def printName(self):
        print(self.emailID, self.fullName)

    def printDetails(self):
        print(self.emailID, self.fullName, self.phoneNum)


def traineeMenu():
    ch = 0

    while(ch!="9"):
        print("Trainee Menu: ")
        print("1: Add Trainee, 2: Update Trainee, 3:Delete Trainee")
        print("4: Add Trainee to a course")
        print("9: End")
        print("Please enter your choice: ")
        ch = input()
        print()

        if (ch == "1"):
            id = input("Please enter trainee's id: ")
            index = getIndex(id, "ListOfTrainees")
            if (index == -1):
                name = input("Please enter trainee's name: ")
                degree = input("Please enter trainee's degree: ")
                workExp = input("Please enter trainee's work exerience: ")
                newTrainee = trainee(id, name, degree, workExp)
                newTrainee.addTrainee()
            else:
                print("Error, ID already in use")

        elif (ch == "2"):
            id = input("Please enter trainee's id: ")
            index = getIndex(id, "ListOfTrainees")
            if(index !=-1):
                ws = wb["ListOfTrainees"]
                currTrainee= trainee(id, ws.cell(row=index,column=2).value, ws.cell(row=index,column=3).value, ws.cell(row=index,column=4).value)
                print("Current trainee:", end=" ")
                currTrainee.printDetails()

                name = input("Please enter trainee's name: ")
                degree = input("Please enter trainee's degree: ")
                workExp = input("Please enter trainee's work exerience: ")
                updatedTrainee = trainee(id, name, degree, workExp)
                updatedTrainee.updateTrainee()
            else:
                print("Error, ID not in use")

        elif (ch == "3"):
            id = input("Please enter trainee's id: ")
            #Delete trainee
            ws = wb["ListOfTrainees"]
            index = getIndex(id,"ListOfTrainees")
            ws.delete_rows(index, 1)
            wb.save("excelProjectDB.xlsx")

            ws = wb["MappingCourseTrainee"]
            for i in range(1, ws.max_row+1):
                if(ws.cell(row=i, column=2).value==id):
                    ws.delete_rows(i, 1)
                    wb.save("excelProjectDB.xlsx")

        elif (ch == "4"):
            traineeID = input("Please enter trainee's id: ")
            index = getIndex(traineeID, "ListOfTrainees")
            if (index == -1):
                print("Error, Trainee doesn't exist")
                continue

            courseID = input("Please enter course's id: ")
            index = getIndex(courseID, "CourseDetails")
            if (index == -1):
                print("Error, Course doesn't exist")
                continue

            ws = wb["MappingCourseTrainee"]
            new_row = (courseID, traineeID)
            ws.append(new_row)
            wb.save("excelProjectDB.xlsx")

        elif (ch == "9"):
            print("Returning to main")

        else:
            print("Invalid input")

        print()

def courseMenu():
    ch = 0

    while (ch != "9"):
        print("Course Menu: ")
        print("1: Add Course, 2: Update Course")
        print("9: End")
        print("Please enter your choice: ")
        ch = input()
        print()

        if (ch == "1"):
            id = input("Please enter Course's id: ")
            index = getIndex(id, "CourseDetails")
            if (index == -1):
                description = input("Please enter Course's description: ")
                newCourse = course(id, description)
                newCourse.addCourse()
            else:
                print("Error, ID already in use")

        elif (ch == "2"):
            id = input("Please enter Course's id: ")
            index = getIndex(id, "CourseDetails")
            if (index != -1):
                ws = wb["CourseDetails"]
                currCourse = course(id, ws.cell(row=index, column=2).value)
                print("Current Course:", end=" ")
                currCourse.printDetails()

                description = input("Please enter Course's description: ")
                updateCourse = course(id, description)
                updateCourse.updateCourse()
            else:
                print("Error, ID not in use")

        elif (ch == "9"):
            print("Returning to main")

        else:
            print("Invalid input")

        print()

def trainerMenu():
    ch = 0

    while (ch != "9"):
        print("Trainer Menu: ")
        print("1: Add Trainer, 2: Update Trainer, 3:Delete Trainer")
        print("4: Add Trainer to a course")
        print("9: End")
        print("Please enter your choice: ")
        ch = input()
        print()

        if (ch == "1"):
            id = input("Please enter Trainer's email: ")
            index = getIndex(id, "TrainerDetails")
            if (index == -1):
                fullName = input("Please enter Trainer's full Name: ")
                phoneNum = input("Please enter Trainer's phone number: ")
                newTrainer = trainer(id, fullName, phoneNum)
                newTrainer.addTrainer()
            else:
                print("Error, Email already in use")

        elif (ch == "2"):
            id = input("Please enter Trainer's id: ")
            index = getIndex(id, "TrainerDetails")
            if (index != -1):
                ws = wb["TrainerDetails"]
                currTrainer = trainer(id, ws.cell(row=index, column=2).value, ws.cell(row=index, column=3).value)
                print("Current trainer:", end=" ")
                currTrainer.printDetails()

                fullName = input("Please enter Trainer's full Name: ")
                phoneNum = input("Please enter Trainer's phone number: ")
                updatedTrainer = trainer(id, fullName, phoneNum)
                updatedTrainer.updateTrainer()
            else:
                print("Error, ID not in use")

        elif (ch == "3"):
            id = input("Please enter Trainer's id: ")
            # Delete trainer
            ws = wb["TrainerDetails"]
            index = getIndex(id, "TrainerDetails")
            ws.delete_rows(index, 1)
            wb.save("excelProjectDB.xlsx")

            ws = wb["MappingCourseTrainer"]
            for i in range(1, ws.max_row + 1):
                if (ws.cell(row=i, column=2).value == id):
                    ws.delete_rows(i, 1)
                    wb.save("excelProjectDB.xlsx")

        elif (ch == "4"):
            trainerID = input("Please enter trainer's id: ")
            index = getIndex(trainerID, "TrainerDetails")
            if (index == -1):
                print("Error, Trainer doesn't exist")
                continue

            courseID = input("Please enter course's id: ")
            index = getIndex(courseID, "CourseDetails")
            if (index == -1):
                print("Error, Course doesn't exist")
                continue

            index = getIndex(courseID, "MappingCourseTrainer")
            if (index != -1):
                print("Error, Course already has a trainer")
                continue

            ws = wb["MappingCourseTrainer"]
            new_row = (courseID, trainerID)
            ws.append(new_row)
            wb.save("excelProjectDB.xlsx")

        elif (ch == "9"):
            print("Returning to main")

        else:
            print("Invalid input")

        print()

def managerMenu():
    ch = 0

    while (ch != "9"):
        print("Manager Menu: ")
        print("1: Add Manager, 2: Update Manager, 3:Delete Manager")
        print("9: End")
        print("Please enter your choice: ")
        ch = input()
        print()

        if (ch == "1"):
            id = input("Please enter Manager's email: ")
            index = getIndex(id, "ManagerDetails")
            if (index == -1):
                fullName = input("Please enter Manager's full Name: ")
                phoneNum = input("Please enter Manager's phone number: ")
                newManager = manager(id, fullName, phoneNum)
                newManager.addManager()
            else:
                print("Error, Email already in use")

        elif (ch == "2"):
            id = input("Please enter Manager's id: ")
            index = getIndex(id, "ManagerDetails")
            if (index != -1):
                ws = wb["ManagerDetails"]
                currManager = trainer(id, ws.cell(row=index, column=2).value, ws.cell(row=index, column=3).value)
                print("Current Manager:", end=" ")
                currManager.printDetails()

                fullName = input("Please enter Manager's full Name: ")
                phoneNum = input("Please enter Manager's phone number: ")
                updatedManager = manager(id, fullName, phoneNum)
                updatedManager.updateManager()
            else:
                print("Error, ID not in use")

        elif (ch == "3"):
            id = input("Please enter Manager's id: ")
            # Delete manager
            ws = wb["ManagerDetails"]
            index = getIndex(id, "ManagerDetails")
            ws.delete_rows(index, 1)
            wb.save("excelProjectDB.xlsx")

        elif (ch == "9"):
            print("Returning to main")

        else:
            print("Invalid input")

        print()

def sessionMenu():
    ch = 0

    while (ch != "9"):
        print("Session Menu: ")
        print("1: Add Session, 2: View Session, 3: Mark Absences")
        print("9: End")
        print("Please enter your choice: ")
        ch = input()
        print()

        if (ch == "1"):
            sessionDate = input("Please enter the date of the session:")
            classID = input("Please enter the id of the class of the session:")
            index = getIndex(classID, "CourseDetails")
            if (index == -1):
                print("Error, class doesn't exist")
                continue

            sheets = wb.sheetnames
            if (classID+" "+sessionDate) in sheets:
                print("Error, class session already created for that date")
                continue

            wb.create_sheet(classID + " " + sessionDate)
            wb.save("excelProjectDB.xlsx")

            startTime = input("Please enter the start time of the session:")
            endTime = input("Please enter the end time of the session:")

            ws = wb[classID+" "+sessionDate]
            new_row = (startTime, endTime)
            ws.append(new_row)
            wb.save("excelProjectDB.xlsx")

            #Load in all students
            wr = wb["MappingCourseTrainee"]
            for i in range(1, ws.max_row+1):
                if(wr.cell(row=i,column=1).value==classID):
                    new_row = (wr.cell(row=i,column=2).value, "Present")
                    ws.append(new_row)
                    wb.save("excelProjectDB.xlsx")

        elif(ch == "2"):
            sessionDate = input("Please enter the date of the session:")
            classID = input("Please enter the id of the class of the session:")
            index = getIndex(classID, "CourseDetails")
            if (index == -1):
                print("Error, class doesn't exist")
                continue

            sheets = wb.sheetnames
            if not((classID + " " + sessionDate) in sheets):
                print("Error, Class session not created for that date")
                continue

            print()
            print("Session:",sessionDate)
            ws = wb["CourseDetails"]
            currCourse = course(classID, ws.cell(row=index, column=2).value)
            print("Class:", end=" ")
            currCourse.printDetails()

            ws = wb[classID + " " + sessionDate]
            print("Start:",ws.cell(row=1,column=1).value,"End:",ws.cell(row=1,column=2).value)


            index= getIndex(classID, "MappingCourseTrainer")
            if(index!=-1):
                ws = wb["MappingCourseTrainer"]
                teachid= ws.cell(row=index, column=2).value
                index = getIndex(teachid, "TrainerDetails")
                ws = wb["TrainerDetails"]
                teacher = trainer(teachid, ws.cell(row=index, column=2).value, ws.cell(row=index, column=3).value)
                print("Trainer:", end=" ")
                teacher.printDetails()

            ws = wb[classID + " " + sessionDate]
            wr = wb["ListOfTrainees"]
            for i in range(2, ws.max_row+1):
                index = getIndex(ws.cell(row=i, column=1).value, "ListOfTrainees")
                studentName = wr.cell(row=index, column=2).value
                print("Student:",studentName,ws.cell(row=i,column=2).value)

        elif (ch == "3"):
            sessionDate = input("Please enter the date of the session:")
            classID = input("Please enter the id of the class of the session:")
            index = getIndex(classID, "CourseDetails")
            if (index == -1):
                print("Error, class doesn't exist")
                continue

            sheets = wb.sheetnames
            if not ((classID + " " + sessionDate) in sheets):
                print("Error, Class sesson not created for that date")
                continue

            #Take list of students, find them and mark absent
            inputVal = input("Please input the list of student IDs: ")
            studentIDList = inputVal.split(" ")

            ws = wb[classID + " " + sessionDate]
            students = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row+1)]

            for x in range (0,len(students)):
                if (students[x] in studentIDList):
                    ws.cell(row=2+x, column=2, value = "Absent")


        elif (ch == "9"):
            print("Returning to main")

        else:
            print("Invalid input")

        print()


#Main
ch = 0

while(ch!="9"):
    print("Main Menu: ")
    print("1: Trainee Menu, 2: Course Menu, 3: Trainer Menu")
    print("4: Manager Menu, 5: Session Menu")
    print("9: End")
    print("Please enter your choice: ")
    ch = input()
    print()

    if(ch=="1"):
        traineeMenu()
    elif (ch=="2"):
        courseMenu()
    elif (ch=="3"):
        trainerMenu()
    elif (ch=="4"):
        managerMenu()
    elif (ch=="5"):
        sessionMenu()
    elif (ch=="9"):
        print("Thank you for using this program")
    else:
        print("Invalid input")


